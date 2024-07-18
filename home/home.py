from flask import Blueprint,render_template, url_for, redirect, request, Response
from home import *
from flask_login import *
from model import *
from login.login import bcrypt
import openpyxl
from datetime import date
from datetime import datetime
import io
import xlwt
import psycopg2
import psycopg2.extras
import warnings

warnings.simplefilter(action='ignore', category=UserWarning)

titolo = "Home"
@home_bp.route('/', methods=['GET', 'POST'])
def home():
    contatti = conn.execute(select(contatto).where(contatto.c.idUtente == current_user.get_id())).fetchall()
    
    """conn.execute(insert(utente).values(
        nome = 'Marco',
        cognome= 'Piazzon',
        email= 'mp@gmail.com',
        psw=bcrypt.generate_password_hash('mp@gmail.com'+'ciao').decode('utf-8'),
        ))
    conn.commit()"""
    getPortafogliUtente = conn.execute(select(portafoglio).where(portafoglio.c.idUtente == current_user.get_id()).order_by(portafoglio.c.idPortafoglio.desc())).fetchall()
    getPortafogliUtente = list(getPortafogliUtente)
    print(getPortafogliUtente)
    print(type(getPortafogliUtente))
    print("ho fatto questo")
    return render_template ("/home/home.html", getP=getPortafogliUtente, len=len(getPortafogliUtente), titolo=titolo)

@home_bp.route("/delete/<int:id>", methods=['POST','GET'])
@login_required
def removePortafoglio(id):
    print("rimuovo")
    print("sto cancellando")
    print(id)
    try:
        #Cancello tabella appuntamento 
        # DELETE appuntamento FROM appuntamento join trattativaappuntamento on appuntamento.idAppuntamento = trattativaappuntamento.idAppuntamento 
        # join trattativa on trattativa.idTrattativa = trattativaappuntamento.idTrattativa join cliente on trattativa.idCliente = cliente.idCliente where cliente.idPortafoglio = 6;
        conn.execute(delete(appuntamento).\
                                    where(appuntamento.c.idAppuntamento == trattativaappuntamento.c.idAppuntamento).\
                                    where(trattativaappuntamento.c.idTrattativa == trattativa.c.idTrattativa).\
                                    where(trattativa.c.idCliente == cliente.c.idCliente).\
                                    where(cliente.c.idPortafoglio == id)                                                    
        )

        #Cancello tabella trattativaappuntamento


        conn.execute(delete(trattativaappuntamento).\
                where(trattativa.c.idCliente == cliente.c.idCliente).\
                where(trattativaappuntamento.c.idTrattativa == trattativa.c.idTrattativa).\
                where(trattativa.c.idCliente == cliente.c.idCliente).\
                where(cliente.c.idPortafoglio == id)
        )
        

        #Cancello trattattive
        # delete trattativa from trattativa join cliente on trattativa.idCliente = cliente.idCliente where cliente.idPortafoglio = 6;
        conn.execute(delete(trattativa).\
                    where(trattativa.c.idCliente == cliente.c.idCliente).\
                    where(cliente.c.idPortafoglio == id)
        )
        
        #Cancello clienti
        conn.execute(delete(cliente).where(cliente.c.idPortafoglio == id))

        #Cancello portafoglio

        conn.execute(delete(portafoglio).where(portafoglio.c.idPortafoglio == id))
        
        conn.commit()
        print("ho cancellato quello che dovevo fare")
    except Exception as error:
        print("rip")
        print(error.__cause__)
        print(error)
        conn.rollback()


    return redirect(url_for('home_bp.home'))


@home_bp.route('/portafoglio/<int:id>')
@login_required
def goToPortafoglio(id, message):
    print("cviao")
    print(id)
    clienti = conn.execute(select(cliente).where(cliente.c.idPortafoglio == id)).fetchall()
    return render_template("/portafoglio/portafoglio.html", clienti=clienti, message = message, titolo ="Portafoglio")

def filterNumber(val):
    if(val is None):
        return None
    
    if(val is String):
        sub = sub.replace(" ", "")
        sub = list(val)
        if(len(sub) == 13 and list(val)[0] == '+'): # esempio '+393453659562' --> 3453659562
            return val[2:]
    
    return str(val)


def getFase(andtra, value):
    print(type(andtra[-1][0]))
    if(value is None):
        return andtra[-1][0]
    for v in andtra:
        if(value.upper() == v.nome.upper()):
            return v.idAndamento
    return andtra[-1][0]

def getCategoria(getCateOff, value):
    print(getCateOff[-1][0])
    if(value is None):
        return getCateOff[-1][0]
    for v in getCateOff:
        if(value.upper() == v.nome.upper()):
            print("trovato" + str(v[0]))
            return v.idCategoria
    return getCateOff[-1][0]
"""
def convertDate(val):
    if(val is None):
        return None
    
    print
    split_val = date.strftime(val, "%").split("-")
    if(len(split_val) != 2):
        return None
     
    months = ["gennaio","febbraio","marzo","aprile","maggio","giugno","luglio","agosto","settembre","ottobre","novembre","dicembre"]
    month = None
    
    for i in range(len(months)):
        if(months[i] == split_val[0]):
            month = i+1
    
    day = str(split_val[1])
    if(month is None or day is None):
        return None
    
    return date(date.year, month, day) 
"""
@home_bp.route('/addPortafoglio',methods=['POST'])
@login_required
def addPortafoglio():
    warnings.simplefilter(action='ignore', category=UserWarning)
    print("sono quya")
    # Read the File using Flask request
    file = request.files['fileDoc']
 
    # Parse the data as a Pandas DataFrame type
    #data = pandas.read_excel(file)
 
    # Define variable to load the dataframe
    dataframe = openpyxl.load_workbook(file)
    
    # Define variable to read sheet
    dataframe1 = dataframe.active

    file2 = request.files['filePipeline']
    # Parse the data as a Pandas DataFrame type
    #data = pandas.read_excel(file)

    newid = 0
    try:
        res = conn.execute(insert(portafoglio).values(
            idUtente = current_user.get_id(),
            dataInserimento = date.today()
        ))

        newid = res.inserted_primary_key[0]
        
        current_user.idPort = newid
        #print("test della vita")
        print("test id")
        print(newid)
        print(res.lastrowid)
        for col in range(1, dataframe1.max_row):
            listapp = []
            for row in dataframe1.iter_cols(1, dataframe1.max_column):
                listapp.append(row[col].value)
           # print("sono qua 3")
           # print(listapp)
            conn.execute(insert(cliente).values(
                idUtente = current_user.get_id(),
                idPortafoglio = newid,
                tipoCliente = conn.execute(select(tipocliente.c.idTipoCliente).where(tipocliente.c.nome == listapp[0])).fetchone()[0], #da testare
                cf = listapp[1],
                ragioneSociale = listapp[2],
                
                presidio = listapp[3],
                indirizzoPrincipale = (listapp[4]),
                comunePrincipale = listapp[5],
                capPrincipale = listapp[6],
                provinciaDescPrincipale = listapp[7], 
                provinciaSiglaPrincipale = listapp[8],
                sediTot = (listapp[9]),
                nLineeTot = (listapp[10]),
                fisso = (listapp[11]),
                mobile = (listapp[12]),
                totale = (listapp[13]),
                fatturatoCerved = (listapp[14]),
                clienteOffMobScadenza = (listapp[15]),
                fatturatoTim = (listapp[16]),
                dipendenti = (listapp[17])
            ))
            print("insert")
        print("okokoko")
        print("sono qua 1")
        if not (file2 is None):
             
            # Define variable to load the dataframe
            dataframe2 = openpyxl.load_workbook(file2, data_only=True)
            
            # Define variable to read sheet
            dataframe3 = dataframe2.active
            #print(dataframe3.max_row)
            andtra = conn.execute(select(andamentotrattativa)).fetchall()
            for m in andtra:
                print(m.idAndamento)
                print(m.nome)
            print(andtra)
            print(andtra[0][0])
            print(type(andtra[0][0]))
            
            getCateOff = conn.execute(select(categoria)).fetchall()
            print("dopo")
            for col in range(1, dataframe3.max_row):
                listapp = []
                checkRow = True
                for row in dataframe3.iter_cols(1, dataframe3.max_column):
                    if(row[col].value is not None):
                        checkRow = False
                    #print((row[col].value), end= ", ")
                    listapp.append(row[col].value)
                
                if(checkRow == False):
                    #print("sto testando id")
                    #print(listapp)
                    #print(listapp[3])
                    idlist = conn.execute(select(cliente.c.idCliente,cliente.c.idPortafoglio).where(listapp[3] == cliente.c.ragioneSociale)).fetchall()
                    print(len(idlist))
                    #print(idlist)
                    idlist = list(idlist)
                    print("lista")
                    print(idlist)
                    numero = idlist[-1]
                    prob = None
                    if(listapp[19]):
                        prob = listapp[19]*100
                    try:
                        conn.execute(insert(trattativa).values(
                            idUtente = current_user.get_id(),
                            idCliente = numero[0], #fare query che trova il nome e assegna l'id nella tabella cliente
                            codiceCtrDigitali = listapp[0],
                            codiceSalesHub = listapp[1],
                            areaManager = conn.execute(select(utente.c.areaManager).where(utente.c.idUtente == current_user.get_id())),
                            zona = listapp[4],
                            tipo = (listapp[5]),
                            nomeOpportunita = listapp[6],
                            dataCreazioneOpportunita = (listapp[7]), # da testare
                            fix = listapp[8],
                            mobile = listapp[9],
                            categoriaOffertaIT = getCategoria(getCateOff, listapp[10]),
                            it = (listapp[11]),
                            lineeFoniaFix = (listapp[12]),
                            aom = (listapp[13]),
                            mnp = (listapp[14]),
                            al = (listapp[15]),
                            dataChiusura = (listapp[16]),
                            fase = getFase(andtra, listapp[17]) ,
                            noteSpecialista = (listapp[18]),
                            probabilita = prob,
                            inPaf = None,
                            record = listapp[21],
                            fornitore = listapp[22],
                        ))
                    except Exception as error:
                        print("rip row interna")
                        print(error)
                        print(error.__cause__)
        print("okokoko")
        global message
        message = "Portafoglio creato"
        conn.commit()
    except Exception as error:
        print("rip")
        print(error)
        print(error.__cause__)
        conn.rollback()
    
    return redirect(url_for("portafoglio_bp.home", idPort = newid, id = 0))


@home_bp.route('/getExcel/<int:id>')
def getExcel(id):
    print("sto creando il file")
    print(id)
    res = conn.execute(select(cliente).where(cliente.c.idPortafoglio == id)).fetchall()

    output = io.BytesIO()

    workbook = xlwt.Workbook()

    sh = workbook.add_sheet('Report Portafoglio')

    sh.write(0,0,'TIPO CLIENTE')
    sh.write(0,1,'CF')
    sh.write(0,2,'RAG_SOC_CLI')
    sh.write(0,3,'PRESIDIO')
    sh.write(0,4,'INDIRIZZO PRINCIPALE')
    sh.write(0,5,'COMUNE_PRINCIPALE')
    sh.write(0,6,'CAP_PRINCIPALE')
    sh.write(0,7,'PROVINCIA_DESCR_PRINCIPALE')
    sh.write(0,8,'PROVINCIA_SIGLA_PRINCIPALE')
    sh.write(0,9,'SEDI_TOT')
    sh.write(0,10,'N_LINEE_TOT')
    sh.write(0,11,'_FISSO')
    sh.write(0,12,'_MOBILE')
    sh.write(0,13,'_TOTALE')
    sh.write(0,14,'FATTURATO_CERVED')
    sh.write(0,15,'CLIENTE_OFF_MOB_SCADENZA')
    sh.write(0,16,'FATTURATO_IT_TIM')
    sh.write(0,17,'DIPENDENTI')

    idx = 0
    for row in res:
        print(row)
        sh.write(idx+1, 0, row['tipoCliente'])
        sh.write(idx+1, 1, row['cf'])
        sh.write(idx+1, 2, row['ragioneSociale'])
        sh.write(idx+1, 3, row['presidio'])
        sh.write(idx+1, 4, row['indirizzoPrincipale'])
        sh.write(idx+1, 5, row['comunePrincipale'])
        sh.write(idx+1, 6, row['capPrincipale'])
        sh.write(idx+1, 7, row['provinciaDescPrincipale'])
        sh.write(idx+1, 8, row['provinciaSiglaPrincipale'])
        sh.write(idx+1, 9, row['sediTot'])
        sh.write(idx+1, 10, row['nLineeTot'])
        sh.write(idx+1, 11, row['fisso'])
        sh.write(idx+1, 12, row['mobile'])
        sh.write(idx+1, 13, row['totale'])
        sh.write(idx+1, 14, row['fatturatoCerved'])
        sh.write(idx+1, 15, row['clienteOffMobScadenza'])
        sh.write(idx+1, 16, row['fatturatoTim'])
        sh.write(idx+1, 17, row['dipendenti'])
        idx += 1
    
    workbook.save(output)
    output.seek(0)
    
    return Response(output, mimetype="application/ms-excel", headers={"Content-Disposition":"attachment;filename=report_portafoglio.xls"})
        
""""
  idUtente = current_user.get_id(),
                idPortafoglio = lastPortafoglio,
                tipoCliente = 1,
                cf = correctValue(list[1],
                ragioneSociale = correctValue(list[2]),
                presidio = 1,
                indirizzoPrincipale = correctValue(list[4]),
                comunePrincipale = 1,
                capPrincipale = correctValue(list[6]),
                provinciaDescPrincipale = correctValue(list[7]),
                provinciaSiglaPrincipale = correctValue(list[8]),
                sediTot = correctValue(list[9]),
                nLineeTot = correctValue(list[10]),
                fisso = correctValue(list[11]),
                mobile = correctValue(list[12]),
                totale = correctValue(list[12]),
                fatturatoCerved = correctValue(list[13]),
                clienteOffMobScadenza = correctValue(list[14]),
                fatturatoTim = correctValue(list[15]),
                dipendenti = correctValue(list[16])"""